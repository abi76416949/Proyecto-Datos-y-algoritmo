from base_de_datos import BaseDeDatos
from persona import Persona
import openpyxl
from openpyxl import Workbook
from prettytable import PrettyTable

#clase para el registro de personas
class RegistroPersona:
    #funcion para inicializar el registro
    def __init__(self):
        self.bd = BaseDeDatos('personas.xlsx')
        self.bd.cerrar()
        datos = self.bd.obtener_datos()
        if isinstance(datos, list) and all(isinstance(d, dict) for d in datos):
            self.personas = [Persona(**persona) for persona in datos]
        else:
            self.personas = []

    #funcion para agregar una persona al registro

    def agregar_persona(self, persona):
        try:
            wb = openpyxl.load_workbook('personas.xlsx')
            sheet = wb.active
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active
            sheet.append(['Nombre', 'Codigo', 'Edad', 'Número', 'Género', 'Fecha de Nacimiento', 'Correo'])
        sheet.append([persona.nombre, persona.codigo, persona.edad, persona.numero, persona.genero, persona.fecha_nacimiento, persona.correo])
        wb.save('personas.xlsx')
        self.personas.append(persona)
        print("Persona agregada exitosamente.")

    #funcion para eliminar una persona del registro
    def eliminar_persona(self, codigo):
        personas_filtradas = [persona for persona in self.personas if persona.codigo != codigo]
        if len(personas_filtradas) < len(self.personas):
            self.personas = personas_filtradas
            self.bd.guardar_datos([persona.__dict__ for persona in self.personas])
            print(f"Persona {codigo} eliminada.")
        else:
            print(f"No se encontró a {codigo} en el registro.")

    #funcion para mostrar las personas del registro
    def mostrar_personas(self):
        table = PrettyTable()
        table.field_names = ["Nombre", "Codigo", "Edad", "Número", "Género", "Fecha de Nacimiento", "Correo"]
        for persona in self.personas:
            table.add_row([persona.nombre, persona.codigo, persona.edad, persona.numero, persona.genero, persona.fecha_nacimiento, persona.correo])
        print(table)

    #funcion para buscar una persona en el registro por codigo
    def buscar_persona_por_codigo(self, codigo):
        personas_encontradas = [persona for persona in self.personas if persona.codigo == codigo]
        if personas_encontradas:
            for persona in personas_encontradas:
                print("Persona encontrada:")
                print(f"Nombre: {persona.nombre}\nCodigo: {persona.codigo}\nEdad: {persona.edad}\nNúmero: {persona.numero}\nGénero: {persona.genero}\nFecha de Nacimiento: {persona.fecha_nacimiento}\nCorreo: {persona.correo}\n")
        else:
            print(f"No se encontró a {codigo} en el registro.")

    #funcion para buscar una persona en el registro por nombre
    def buscar_persona_por_nombre(self, nombre):
        wb = openpyxl.load_workbook('personas.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            if row[0] == nombre:
                print(f"Nombre: {row[0]}\nCodigo: {row[1]}\nEdad: {row[2]}\nNúmero: {row[3]}\nGénero: {row[4]}\nFecha de Nacimiento: {row[5]}\nCorreo: {row[6]}\n")
                break
        else:
            print(f"No se encontró a {nombre} en el registro.")


    # funcion para editar una persona del registro por codigo
    def editar_persona_por_codigo(self, codigo):
        personas_encontradas = [persona for persona in self.personas if persona.codigo == codigo]
        if personas_encontradas:
            for persona in personas_encontradas:
                print("Persona encontrada:")
                print(f"Nombre: {persona.nombre}\nCodigo: {persona.codigo}\nEdad: {persona.edad}\nNúmero: {persona.numero}\nGénero: {persona.genero}\nFecha de Nacimiento: {persona.fecha_nacimiento}\nCorreo: {persona.correo}\n")
                print("Ingrese los nuevos datos de la persona:")
                nombre = input("Nombre y apellido: ")
                edad = int(input("Edad: "))
                codigo = input("Codigo: ")
                numero = input("Número de teléfono: ")
                genero = input("Genero(F/M): ")
                fecha_nacimiento = input("Fecha de nacimiento(dd/mm/aaaa): ")
                correo = input("Correo electrónico: ")
                persona.nombre = nombre
                persona.edad = edad
                persona.codigo = codigo
                persona.numero = numero
                persona.genero = genero
                persona.fecha_nacimiento = fecha_nacimiento
                persona.correo = correo
                self.bd.guardar_datos([persona.__dict__ for persona in self.personas])
                print("Persona editada exitosamente.")
                break
        else:
            print(f"No se encontró a {codigo} en el registro.")
